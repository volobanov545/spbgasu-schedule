#!/usr/bin/env python3
"""
Парсер расписания СПбГАСУ → .ics
Источник: rasp.spbgasu.ru (группа 3-СУЗСс-2)

Стратегия получения данных (в порядке приоритета):
  1. Excel endpoint (requests, без браузера)
  2. HTML страница через Playwright
  3. Fallback: saved_resource.html из репозитория
"""

import re
import sys
import hashlib
import argparse
from datetime import datetime, date, timedelta
from pathlib import Path
from io import BytesIO

import requests
from bs4 import BeautifulSoup
from icalendar import Calendar, Event
from icalendar.prop import vText

GROUP = "3-СУЗСс-2"  # последняя цифра = номер курса (2→3→4 каждый сентябрь)
GROUP_ENCODED = "3-%D0%A1%D0%A3%D0%97%D0%A1%D1%81-2"
SCHEDULE_URL = "https://rasp.spbgasu.ru/"
EXCEL_URL = f"https://rasp.spbgasu.ru/getExcel.php?TYPE=GROUPS&FIND={GROUP_ENCODED}"
FALLBACK_HTML = Path(__file__).parent / "saved_resource.html"
OUTPUT_FILE = Path(__file__).parent / "schedule.ics"
SESSION_OUTPUT_FILE = Path(__file__).parent / "session.ics"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/120.0 Safari/537.36",
    "Referer": "https://www.spbgasu.ru/",
}


# ─── Получение данных ────────────────────────────────────────────────────────

def fetch_excel() -> list[dict] | None:
    """Скачивает Excel и парсит уроки. Возвращает None при ошибке."""
    try:
        import openpyxl
    except ImportError:
        print("[INFO] openpyxl не установлен, пропускаю Excel-метод")
        return None

    try:
        print(f"[INFO] Пробую Excel endpoint: {EXCEL_URL}")
        resp = requests.get(EXCEL_URL, headers=HEADERS, timeout=20)
        resp.raise_for_status()
        if "html" in resp.headers.get("content-type", "").lower():
            print("[INFO] Excel endpoint вернул HTML (редирект/блокировка)")
            return None

        wb = openpyxl.load_workbook(BytesIO(resp.content), data_only=True)
        lessons = parse_excel(wb)
        if lessons:
            print(f"[OK] Excel: найдено уроков: {len(lessons)}")
        return lessons or None
    except Exception as e:
        print(f"[INFO] Excel endpoint: {e}")
        return None


def fetch_html_requests() -> str | None:
    """Лёгкая попытка без браузера."""
    try:
        resp = requests.get(SCHEDULE_URL, headers=HEADERS, timeout=20)
        resp.raise_for_status()
        if GROUP not in resp.text:
            print("[INFO] requests: данных группы нет (нужен JS-рендеринг)")
            return None
        print(f"[OK] requests: {len(resp.text)} байт")
        return resp.text
    except Exception as e:
        print(f"[INFO] requests: {e}")
        return None


def fetch_html_playwright() -> tuple[str | None, str | None]:
    """JS-рендеринг через Playwright. Возвращает (html_расписание, html_сессия)."""
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("[INFO] playwright не установлен")
        return None, None

    try:
        print("[INFO] Playwright: запускаю браузер...")
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page(extra_http_headers=HEADERS)
            page.goto(SCHEDULE_URL, wait_until="domcontentloaded", timeout=90000)
            page.wait_for_selector(".get_data", timeout=60000)

            group_el = page.locator(f".get_data[data-search='{GROUP}']")
            if group_el.count():
                group_el.first.click()
                page.wait_for_selector(".lesson", timeout=60000)
            else:
                search = page.locator("input[data-var='GROUPS']")
                if search.count():
                    search.fill(GROUP)
                    page.wait_for_timeout(1000)
                    result = page.locator(f".get_data[data-search='{GROUP}']")
                    if result.count():
                        result.first.click()
                        page.wait_for_selector(".lesson", timeout=60000)

            html = page.content()

            # Проверяем наличие вкладки Сессия
            session_html = None
            session_tab = page.locator("#pills-tab .nav-link").filter(
                has_text=re.compile(r"сессия", re.IGNORECASE)
            )
            if session_tab.count():
                print("[INFO] Playwright: обнаружена вкладка Сессия, загружаю...")
                session_tab.first.click()
                page.wait_for_timeout(3000)
                session_html = page.content()
                print("[OK] Playwright: данные сессии получены")

            browser.close()

        if GROUP not in html:
            print("[INFO] Playwright: данные группы не найдены")
            return None, None
        print(f"[OK] Playwright: {len(html)} байт")
        return html, session_html
    except Exception as e:
        print(f"[INFO] Playwright: {e}")
        return None, None


def load_fallback_html() -> str | None:
    if FALLBACK_HTML.exists():
        print(f"[INFO] Использую fallback HTML: {FALLBACK_HTML}")
        return FALLBACK_HTML.read_text(encoding="utf-8")
    print(f"[ERROR] Fallback HTML не найден: {FALLBACK_HTML}")
    return None


def has_session_tab(html: str) -> bool:
    """Проверяет наличие вкладки Сессия в навигации."""
    soup = BeautifulSoup(html, "html.parser")
    for tab in soup.select("#pills-tab .nav-link"):
        if re.search(r"сессия", tab.get_text(), re.IGNORECASE):
            return True
    return False


# ─── Парсинг Excel ───────────────────────────────────────────────────────────

def parse_excel(wb) -> list[dict]:
    """Парсит уроки из Excel-файла расписания СПбГАСУ."""
    lessons = []
    ws = wb.active

    date_col, time_col, subj_col, room_col, teacher_col = None, None, None, None, None

    # Ищем заголовки
    for row in ws.iter_rows(max_row=5):
        for cell in row:
            v = str(cell.value or "").strip().lower()
            if "дата" in v:
                date_col = cell.column
            elif "время" in v:
                time_col = cell.column
            elif "дисципл" in v or "предмет" in v:
                subj_col = cell.column
            elif "ауд" in v or "каб" in v:
                room_col = cell.column
            elif "препод" in v or "фио" in v:
                teacher_col = cell.column

    if not subj_col:
        return []

    current_date = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Дата
        if date_col and row[date_col - 1]:
            val = row[date_col - 1]
            if isinstance(val, datetime):
                current_date = val.date()
            elif isinstance(val, date):
                current_date = val
            else:
                d = parse_date(str(val))
                if d:
                    current_date = d

        if not current_date:
            continue

        subj = str(row[subj_col - 1] or "").strip()
        if not subj:
            continue

        time_str = str(row[time_col - 1] or "") if time_col else ""
        start_t, end_t = parse_time(time_str)
        if not start_t:
            continue

        room = str(row[room_col - 1] or "").strip() if room_col else ""
        teacher = str(row[teacher_col - 1] or "").strip() if teacher_col else ""

        lessons.append({
            "date": current_date,
            "start": start_t,
            "end": end_t,
            "pair": "",
            "subject": subj,
            "room": room,
            "teacher": teacher,
            "week_info": "",
        })

    return lessons


# ─── Парсинг HTML ────────────────────────────────────────────────────────────

def parse_time(text: str):
    m = re.search(r"(\d{1,2}:\d{2})\s*[-–]\s*(\d{1,2}:\d{2})", text)
    if not m:
        return None, None
    return (
        datetime.strptime(m.group(1), "%H:%M").time(),
        datetime.strptime(m.group(2), "%H:%M").time(),
    )


def parse_date(date_str: str) -> date | None:
    try:
        return datetime.strptime(date_str.strip(), "%d.%m.%Y").date()
    except ValueError:
        return None


def clean(text: str) -> str:
    return " ".join(text.split())


def parse_html(html: str) -> list[dict]:
    soup = BeautifulSoup(html, "html.parser")
    lessons = []

    for week_item in soup.select(".item"):
        time_div = week_item.select_one(".time")
        week_info = clean(time_div.get_text()) if time_div else ""

        for day_div in week_item.select(".days"):
            day_name_div = day_div.select_one(".week_day")
            if not day_name_div:
                continue
            date_div = day_name_div.select_one(".date")
            if not date_div:
                continue
            lesson_date = parse_date(date_div.get_text())
            if not lesson_date:
                continue

            for lesson in day_div.select(".lesson"):
                day_name_block = lesson.select_one(".day_name")
                if not day_name_block:
                    continue

                pair_tag = day_name_block.select_one("b")
                pair_text = clean(pair_tag.get_text()) if pair_tag else ""
                start_t, end_t = parse_time(day_name_block.get_text())
                if not start_t:
                    continue

                block = lesson.select_one(".lesson_block")
                if not block:
                    continue

                divs = [clean(d.get_text()) for d in block.find_all("div", recursive=False)]
                if len(divs) < 2:
                    continue

                subject = divs[0]
                room = teacher = ""
                if len(divs) == 4:
                    room, teacher = divs[2], divs[3]
                elif len(divs) == 3:
                    room, teacher = divs[1], divs[2]
                elif len(divs) == 2:
                    room = divs[1]

                lessons.append({
                    "date": lesson_date,
                    "start": start_t,
                    "end": end_t,
                    "pair": pair_text,
                    "subject": subject,
                    "room": room,
                    "teacher": teacher,
                    "week_info": week_info,
                })

    return lessons


def parse_session_html(html: str) -> list[dict]:
    """Парсит экзамены из вкладки Сессия (#pills-S)."""
    soup = BeautifulSoup(html, "html.parser")

    # Ищем блок сессии
    session_block = soup.select_one("#pills-S")
    if not session_block:
        # Если вкладка уже активна — парсим всю страницу как сессионную
        session_block = soup

    lessons = []

    # Пробуем ту же структуру что и у обычного расписания
    for week_item in session_block.select(".item"):
        time_div = week_item.select_one(".time")
        week_info = clean(time_div.get_text()) if time_div else ""

        for day_div in week_item.select(".days"):
            day_name_div = day_div.select_one(".week_day")
            if not day_name_div:
                continue
            date_div = day_name_div.select_one(".date")
            if not date_div:
                continue
            lesson_date = parse_date(date_div.get_text())
            if not lesson_date:
                continue

            for lesson in day_div.select(".lesson"):
                day_name_block = lesson.select_one(".day_name")
                if not day_name_block:
                    continue

                pair_tag = day_name_block.select_one("b")
                pair_text = clean(pair_tag.get_text()) if pair_tag else ""
                start_t, end_t = parse_time(day_name_block.get_text())
                if not start_t:
                    continue

                block = lesson.select_one(".lesson_block")
                if not block:
                    continue

                divs = [clean(d.get_text()) for d in block.find_all("div", recursive=False)]
                if len(divs) < 1:
                    continue

                subject = divs[0]
                room = teacher = ""
                if len(divs) == 4:
                    room, teacher = divs[2], divs[3]
                elif len(divs) == 3:
                    room, teacher = divs[1], divs[2]
                elif len(divs) == 2:
                    room = divs[1]

                lessons.append({
                    "date": lesson_date,
                    "start": start_t,
                    "end": end_t,
                    "pair": pair_text,
                    "subject": subject,
                    "room": room,
                    "teacher": teacher,
                    "week_info": week_info,
                })

    return lessons


# ─── Генерация ICS ───────────────────────────────────────────────────────────

def make_uid(lesson: dict) -> str:
    key = f"{lesson['date']}{lesson['start']}{lesson['subject']}{lesson['room']}"
    return hashlib.md5(key.encode()).hexdigest() + "@spbgasu"


def build_ics(lessons: list[dict], days_ahead: int | None = None) -> tuple[Calendar, int]:
    today = date.today()
    cutoff = today + timedelta(days=days_ahead) if days_ahead else None
    # Начало текущей недели (понедельник) минус 7 дней
    monday = today - timedelta(days=today.weekday())
    start = monday - timedelta(days=7)
    # Конец через 3 полные недели от текущего понедельника (числитель→знаменатель→числитель)
    end = monday + timedelta(weeks=3) - timedelta(days=1)
    if cutoff:
        end = min(end, cutoff)
    filtered = [l for l in lessons if l["date"] >= start and l["date"] <= end]

    cal = Calendar()
    cal.add("prodid", "-//СПбГАСУ Schedule//3-СУЗСс-2//RU")
    cal.add("version", "2.0")
    cal.add("calscale", "GREGORIAN")
    cal.add("x-wr-calname", f"СПбГАСУ {GROUP}")
    cal.add("x-wr-timezone", "Europe/Moscow")
    cal.add("x-wr-caldesc", f"Расписание. Обновлено {today.strftime('%d.%m.%Y')}")

    for lesson in filtered:
        event = Event()
        event.add("summary", lesson["subject"])
        event.add("dtstart", datetime.combine(lesson["date"], lesson["start"]))
        event.add("dtend", datetime.combine(lesson["date"], lesson["end"]))
        event.add("location", vText(lesson["room"]))

        desc_parts = []
        if lesson["teacher"]:
            desc_parts.append(f"Преподаватель: {lesson['teacher']}")
        if lesson["pair"]:
            desc_parts.append(lesson["pair"])
        if lesson["week_info"]:
            desc_parts.append(lesson["week_info"])
        event.add("description", vText("\n".join(desc_parts)))
        event.add("uid", make_uid(lesson))
        cal.add_component(event)

    return cal, len(filtered)


def build_session_ics(lessons: list[dict]) -> tuple[Calendar, int]:
    """Строит ICS для сессии — все экзамены начиная с сегодня."""
    today = date.today()
    filtered = sorted(
        [l for l in lessons if l["date"] >= today],
        key=lambda l: (l["date"], l["start"])
    )

    cal = Calendar()
    cal.add("prodid", "-//СПбГАСУ Session//3-СУЗСс-2//RU")
    cal.add("version", "2.0")
    cal.add("calscale", "GREGORIAN")
    cal.add("x-wr-calname", f"СПбГАСУ {GROUP} — Сессия")
    cal.add("x-wr-timezone", "Europe/Moscow")
    cal.add("x-wr-caldesc", f"Экзамены. Обновлено {today.strftime('%d.%m.%Y')}")

    for lesson in filtered:
        event = Event()
        event.add("summary", f"📝 {lesson['subject']}")
        event.add("dtstart", datetime.combine(lesson["date"], lesson["start"]))
        event.add("dtend", datetime.combine(lesson["date"], lesson["end"]))
        event.add("location", vText(lesson["room"]))

        desc_parts = []
        if lesson["teacher"]:
            desc_parts.append(f"Преподаватель: {lesson['teacher']}")
        if lesson["pair"]:
            desc_parts.append(lesson["pair"])
        event.add("description", vText("\n".join(desc_parts)))
        event.add("uid", make_uid(lesson) + "-session")
        cal.add_component(event)

    return cal, len(filtered)


# ─── Точка входа ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--playwright", action="store_true")
    parser.add_argument("--file", type=Path)
    parser.add_argument("--days", type=int, default=None)
    parser.add_argument("--output", type=Path, default=OUTPUT_FILE)
    parser.add_argument("--session-output", type=Path, default=SESSION_OUTPUT_FILE)
    args = parser.parse_args()

    lessons = None
    session_html = None

    if args.file:
        html = args.file.read_text(encoding="utf-8")
        print(f"[INFO] Читаю из файла: {args.file}")
        lessons = parse_html(html)
        if has_session_tab(html):
            print("[INFO] В файле обнаружена вкладка Сессия")
            session_html = html
    elif args.playwright:
        html, session_html = fetch_html_playwright()
        if html:
            lessons = parse_html(html)
        if not lessons:
            html = load_fallback_html()
            if html:
                lessons = parse_html(html)
                print("[WARN] Playwright не сработал — использую fallback HTML")
    else:
        # Автоматический режим: Excel → HTML requests → Playwright → fallback
        lessons = fetch_excel()

        if not lessons:
            html = fetch_html_requests()
            if html:
                lessons = parse_html(html)
                if has_session_tab(html):
                    print("[INFO] requests: обнаружена вкладка Сессия, запускаю Playwright для её загрузки")
                    _, session_html = fetch_html_playwright()

        if not lessons:
            html, session_html = fetch_html_playwright()
            if html:
                lessons = parse_html(html)

        if not lessons:
            html = load_fallback_html()
            if html:
                lessons = parse_html(html)
                print("[WARN] Использую устаревший fallback HTML — расписание может быть неактуальным")

    if not lessons:
        print("[ERROR] Не удалось получить расписание ни одним методом.")
        sys.exit(1)

    print(f"[INFO] Всего уроков в источнике: {len(lessons)}")
    cal, count = build_ics(lessons, days_ahead=args.days)
    args.output.write_bytes(cal.to_ical())

    period = f"ближайшие {args.days} дней" if args.days else f"{lessons[0]['date']} — {lessons[-1]['date']}"
    print(f"[OK] Событий в .ics: {count} ({period})")
    print(f"[OK] Сохранено: {args.output}")

    # Сессия
    if session_html:
        session_lessons = parse_session_html(session_html)
        if session_lessons:
            session_cal, session_count = build_session_ics(session_lessons)
            args.session_output.write_bytes(session_cal.to_ical())
            print(f"[OK] Сессия: {session_count} экзаменов → {args.session_output}")
        else:
            print("[INFO] Вкладка Сессия есть, но экзаменов не найдено (возможно структура отличается)")
    else:
        print("[INFO] Вкладка Сессия не обнаружена — session.ics не обновлялся")


if __name__ == "__main__":
    main()
